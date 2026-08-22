"""Single combined report: one row per company-year, one column per model."""

import logging
import os
import pandas as pd
import config
from result import Result, R

log = logging.getLogger(__name__)

COLUMNS = [
    "Company name",
    "Sector Name",
    "Financial Year",
    "Market Price",
    "(DCF) Intrinsic Value",
    "(P/E Relative Valuation) Intrinsic Value",
    "(EV/EBITDA Valuation) Intrinsic Value",
    "(XGBoost) Intrinsic Value",
    "(Random Forest) Intrinsic Value",
    "(LightGBM) Intrinsic Value",
]

VALUE_KEYS = [
    ("price", COLUMNS[3]),
    ("dcf", COLUMNS[4]),
    ("pe", COLUMNS[5]),
    ("ev", COLUMNS[6]),
    ("xgboost", COLUMNS[7]),
    ("random_forest", COLUMNS[8]),
    ("lightgbm", COLUMNS[9]),
]


def _cell(r, key):
    res = r.get(key)
    if not isinstance(res, Result):
        res = Result.bad(R.NOT_COMPUTABLE, "value was never computed")
    return res.text()


def build_dataframe(rows):
    out, prev = [], None
    for r in rows:
        if prev is not None and r["company"] != prev:
            out.append({c: "" for c in COLUMNS})  # blank separator row
        rec = {COLUMNS[0]: r["company"], COLUMNS[1]: r["sector"], COLUMNS[2]: r["fy"]}
        for key, col in VALUE_KEYS:
            rec[col] = _cell(r, key)
        out.append(rec)
        prev = r["company"]
    return pd.DataFrame(out, columns=COLUMNS)


def export(rows, basename=None):
    """-> (csv_path, xlsx_path)"""
    basename = basename or config.OUTPUT_BASENAME
    try:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        df = build_dataframe(rows)
        csv_path = os.path.join(config.OUTPUT_DIR, f"{basename}.csv")
        xlsx_path = os.path.join(config.OUTPUT_DIR, f"{basename}.xlsx")

        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        try:
            from openpyxl.styles import Alignment, Font

            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
                sheet = "Intrinsic Values"
                df.to_excel(xw, index=False, sheet_name=sheet)
                ws = xw.sheets[sheet]

                widths = {
                    0: 28,
                    1: 20,
                    2: 14,
                    3: 26,
                    4: 34,
                    5: 34,
                    6: 34,
                    7: 34,
                    8: 34,
                    9: 34,
                }
                for i, col in enumerate(COLUMNS, start=1):
                    letter = ws.cell(row=1, column=i).column_letter
                    ws.column_dimensions[letter].width = widths.get(i - 1, 20)

                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="center", horizontal="center"
                    )
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.freeze_panes = "D2"  # keep name/sector/FY visible while scrolling
        except Exception as e:
            log.exception("xlsx write failed")
            xlsx_path = f"(xlsx skipped: {type(e).__name__}: {e})"

        return csv_path, xlsx_path
    except Exception as e:
        log.exception("export failed")
        return f"(export failed: {type(e).__name__}: {e})", ""
