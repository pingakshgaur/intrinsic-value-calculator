# FILE: src/io_layer/exporter.py
"""
Four outputs:
    Intrinsic_Value_Report.*        clean grid - values (* = estimated) or a code
    Intrinsic_Value_Data.csv        numeric twin of the grid, for the analyzer
    Intrinsic_Value_Diagnostics.*   full explanation for every empty cell
    Intrinsic_Value_Methods.*       provenance for every filled cell

The grid is 16 columns: the original 10 plus one difference column immediately
after each model's intrinsic value.

    difference = Market Price - Intrinsic Value

    positive -> market trades ABOVE the model's estimate (model says overvalued)
    negative -> market trades BELOW the model's estimate (model says undervalued)

State that convention in the thesis. The opposite one flips the sign of every
bias statistic downstream.

"Market Price" is the 1-MAY BENCHMARK (d["price_bench"]): the close on the
trading day nearest 1 May of the calendar year in which the FY ends. FY2023-24
is scored against the close nearest 1-May-2024. The in-FY valuation anchor is
still carried, in the numeric twin only, as "Valuation Anchor Price" so the two
can be compared during review.

The numeric twin also carries "Benchmark Date", the date that close was
actually taken from. It is rarely 1 May itself - that is Maharashtra Day and
the exchanges are shut - so publishing the realised date is what makes the
benchmark auditable rather than merely asserted.
"""

import logging

import pandas as pd

import config
from result import Result, R
from fy_utils import fy_label

log = logging.getLogger(__name__)

# (row key, intrinsic-value column, difference column)
MODEL_COLUMNS = [
    ("dcf", "(DCF) Intrinsic Value", "Difference (Market - DCF)"),
    (
        "pe",
        "(P/E Relative Valuation) Intrinsic Value",
        "Difference (Market - P/E Rel. Val.)",
    ),
    ("ev", "(EV/EBITDA Valuation) Intrinsic Value", "Difference (Market - EV/EBITDA)"),
    ("xgboost", "(XGBoost) Intrinsic Value", "Difference (Market - XGBoost)"),
    (
        "random_forest",
        "(Random Forest) Intrinsic Value",
        "Difference (Market - Random Forest)",
    ),
    ("lightgbm", "(LightGBM) Intrinsic Value", "Difference (Market - LightGBM)"),
]

BASE_COLUMNS = ["Company name", "Sector Name", "Financial Year", "Market Price"]

COLUMNS = BASE_COLUMNS + [
    c for _, iv_col, diff_col in MODEL_COLUMNS for c in (iv_col, diff_col)
]

# kept for the diagnostics / methods sheets, which iterate every value cell
VALUE_KEYS = [("price", COLUMNS[3])] + [
    (key, iv_col) for key, iv_col, _ in MODEL_COLUMNS
]

DIAG_COLUMNS = [
    "Company name",
    "Sector Name",
    "Financial Year",
    "Model",
    "Reason code",
    "Explanation",
]

METHOD_COLUMNS = [
    "Company name",
    "Financial Year",
    "Model",
    "Value",
    "Estimated?",
    "Method",
]

DATA_COLUMNS = COLUMNS + [
    "FY (internal)",
    "Benchmark Date",
    "Valuation Anchor Price",
]


def _res(r, key):
    v = r.get(key)
    return (
        v
        if isinstance(v, Result)
        else Result.bad(R.NOT_COMPUTABLE, "value was never computed")
    )


def difference(price_res: Result, model_res: Result) -> Result:
    """
    Market Price - Intrinsic Value, as a Result.

    Left blank whenever either side is unavailable. A missing valuation must
    never become a zero difference: zero reads as a perfect prediction and
    would silently corrupt every statistic the analyzer computes. The adjacent
    columns already carry the reason, so repeating the code here would only
    double the noise in the sheet.
    """
    if not price_res.ok or not model_res.ok:
        return Result.bad(R.NOT_COMPUTABLE, "")
    method = None
    if price_res.estimated or model_res.estimated:
        method = "derived from at least one estimated input"
    return Result.good(price_res.value - model_res.value, method=method)


def _diff_text(res: Result) -> str:
    """Blank rather than a reason code, per the note in difference()."""
    return res.text() if res.ok else ""


def build_dataframe(rows):
    out, prev = [], None
    for r in rows:
        if prev is not None and r["company"] != prev:
            out.append({c: "" for c in COLUMNS})

        price = _res(r, "price")
        rec = {
            COLUMNS[0]: r["company"],
            COLUMNS[1]: r["sector"],
            COLUMNS[2]: fy_label(r["fy"]),
            COLUMNS[3]: price.text(),
        }
        for key, iv_col, diff_col in MODEL_COLUMNS:
            model = _res(r, key)
            rec[iv_col] = model.text()
            rec[diff_col] = _diff_text(difference(price, model))
        out.append(rec)
        prev = r["company"]
    return pd.DataFrame(out, columns=COLUMNS)


def build_data_frame(rows):
    """
    Numeric twin: identical headers, plain floats, no blank separator rows, no
    currency glyphs, no reason codes. The analyzer reads this.
    """
    out = []
    for r in rows:
        price = _res(r, "price")
        anchor = _res(r, "price_anchor")
        rec = {
            COLUMNS[0]: r["company"],
            COLUMNS[1]: r["sector"],
            COLUMNS[2]: fy_label(r["fy"]),
            COLUMNS[3]: round(price.value, 4) if price.ok else None,
            "FY (internal)": r["fy"],
            "Benchmark Date": r.get("price_bench_date") or "",
            "Valuation Anchor Price": round(anchor.value, 4) if anchor.ok else None,
        }
        for key, iv_col, diff_col in MODEL_COLUMNS:
            model = _res(r, key)
            d = difference(price, model)
            rec[iv_col] = round(model.value, 4) if model.ok else None
            rec[diff_col] = round(d.value, 4) if d.ok else None
        out.append(rec)
    return pd.DataFrame(out, columns=DATA_COLUMNS)


def build_diagnostics(rows):
    recs = []
    for r in rows:
        for key, col in VALUE_KEYS:
            res = _res(r, key)
            if res.ok:
                continue
            recs.append(
                {
                    DIAG_COLUMNS[0]: r["company"],
                    DIAG_COLUMNS[1]: r["sector"],
                    DIAG_COLUMNS[2]: fy_label(r["fy"]),
                    DIAG_COLUMNS[3]: col,
                    DIAG_COLUMNS[4]: res.code or "",
                    DIAG_COLUMNS[5]: res.detail or "",
                }
            )
    return pd.DataFrame(recs, columns=DIAG_COLUMNS)


def build_methods(rows):
    recs = []
    for r in rows:
        for key, col in VALUE_KEYS:
            res = _res(r, key)
            if not res.ok:
                continue
            method = res.method or "directly computed from reported figures"
            # For the benchmark row, name the realised date. Without it the
            # Methods sheet asserts a 1 May price without ever showing which
            # session supplied it.
            if key == "price" and r.get("price_bench_date"):
                method = (
                    f"close on {r['price_bench_date']} (nearest to 1 May); {method}"
                )
            recs.append(
                {
                    METHOD_COLUMNS[0]: r["company"],
                    METHOD_COLUMNS[1]: fy_label(r["fy"]),
                    METHOD_COLUMNS[2]: col,
                    METHOD_COLUMNS[3]: round(res.value, 2),
                    METHOD_COLUMNS[4]: "YES" if res.estimated else "no",
                    METHOD_COLUMNS[5]: method,
                }
            )
    return pd.DataFrame(recs, columns=METHOD_COLUMNS)


def _style(ws, df, widths, wrap_last=False):
    from openpyxl.styles import Alignment, Font

    for i, col in enumerate(df.columns, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = widths.get(i - 1, 20)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )
    last = len(df.columns)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            wrap = wrap_last and cell.column == last
            cell.alignment = Alignment(
                wrap_text=wrap,
                vertical="top",
                horizontal="right" if cell.column > 3 else "left",
            )


def _report_widths():
    w = {0: 34, 1: 26, 2: 14, 3: 16}
    for i in range(4, len(COLUMNS)):
        # difference columns sit at odd offsets from 5 onwards
        w[i] = 22 if (i - 4) % 2 == 0 else 20
    return w


def export(rows):
    try:
        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        df = build_dataframe(rows)
        csv_path = config.OUTPUT_DIR / f"{config.OUTPUT_BASENAME}.csv"
        xlsx_path = config.OUTPUT_DIR / f"{config.OUTPUT_BASENAME}.xlsx"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        aux_paths = []
        dd = mm = None

        data_df = build_data_frame(rows)
        data_path = config.OUTPUT_DIR / f"{config.DATA_BASENAME}.csv"
        data_df.to_csv(data_path, index=False, encoding="utf-8-sig")
        aux_paths.append(data_path)

        if config.WRITE_DIAGNOSTICS:
            dd = build_diagnostics(rows)
            p = config.OUTPUT_DIR / f"{config.DIAGNOSTICS_BASENAME}.csv"
            dd.to_csv(p, index=False, encoding="utf-8-sig")
            aux_paths.append(p)

        if getattr(config, "WRITE_METHOD_SHEET", False):
            mm = build_methods(rows)
            p = config.OUTPUT_DIR / f"{config.METHODS_BASENAME}.csv"
            mm.to_csv(p, index=False, encoding="utf-8-sig")
            aux_paths.append(p)

        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
                df.to_excel(xw, index=False, sheet_name="Intrinsic Values")
                _style(xw.sheets["Intrinsic Values"], df, _report_widths())
                xw.sheets["Intrinsic Values"].freeze_panes = "D2"

                if dd is not None:
                    dd.to_excel(xw, index=False, sheet_name="Diagnostics")
                    _style(
                        xw.sheets["Diagnostics"],
                        dd,
                        {0: 34, 1: 26, 2: 14, 3: 30, 4: 26, 5: 90},
                        wrap_last=True,
                    )
                    xw.sheets["Diagnostics"].freeze_panes = "A2"

                if mm is not None:
                    mm.to_excel(xw, index=False, sheet_name="Methods")
                    _style(
                        xw.sheets["Methods"],
                        mm,
                        {0: 34, 1: 14, 2: 30, 3: 16, 4: 12, 5: 80},
                        wrap_last=True,
                    )
                    xw.sheets["Methods"].freeze_panes = "A2"
        except Exception as e:
            log.exception("xlsx write failed")
            xlsx_path = f"(xlsx skipped: {type(e).__name__}: {e})"

        return csv_path, xlsx_path, aux_paths
    except Exception as e:
        log.exception("export failed")
        return f"(export failed: {type(e).__name__}: {e})", "", []
